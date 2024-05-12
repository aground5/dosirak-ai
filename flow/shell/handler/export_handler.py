from ...job import dosirak_job
from ...tools import process_date
from ...vo import FlowUser

dosirak_type = ['B', 'A', 'C']
LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

def excel_style(row, col):
    """ Convert given row and column number to an Excel-style cell name. """
    result = []
    while col:
        col, rem = divmod(col-1, 26)
        result[:0] = LETTERS[rem]
    return ''.join(result) + str(row)


async def export_handler(user: FlowUser, input_array: list):
    if len(input_array) == 1:
        from datetime import date
        month = date.today().month
        orders = await dosirak_job.export_month_order(user, month)
        formatted = export_format_orders(orders)
        export_formatted_to_excel(formatted, 3)
    else:
        pass


def export_format_orders(exports):
    formatted = [{} for i in range(31)]
    for export in exports:
        post = export["post"]
        order = export["order"]
        post_title = post["COMMT_TTL"]
        post_date = process_date.parse_date_from_title(post_title)
        formatted_day = formatted[post_date.day - 1]
        for i in range(3):
            order_by_type = order[i]
            dtype = dosirak_type[i]
            for person in order_by_type:
                full_name = f"{person[0]} {person[1]}"
                try:
                    formatted_day[full_name] += dtype * person[2]
                except KeyError:
                    formatted_day[full_name] = dtype * person[2]
    return formatted


def export_formatted_to_excel(formatted: [dict], month: int):
    from openpyxl import load_workbook
    wb = load_workbook("도시락 정산 서식.xltx")
    wb.template = False
    ws = wb.get_sheet_by_name("Sheet1")

    ws['A4'] = f'{month}월'

    people_names = set()
    for day in formatted:
        names = day.keys()
        people_names.update(names)
    for idx, name in enumerate(people_names):
        ws[f'A{5 + idx}'] = name

    name_list = list(people_names)

    for day, day_d in enumerate(formatted):
        for name, value in day_d.items():
            idx = name_list.index(name)
            ws[excel_style(5 + idx, 2 + day)] = value

    wb.save("도시락 정산 출력.xlsx")